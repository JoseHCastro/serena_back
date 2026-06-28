"""Reports service — PDF and Excel generation for clinical reports."""

import uuid
from datetime import datetime, date
from io import BytesIO

from loguru import logger
from sqlalchemy import func, select, and_
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.exceptions import NotFoundError
from app.modules.biometric.repository import EmotionalSnapshotRepository, MicroexpressionRepository
from app.modules.patients.models import Patient
from app.modules.patients.repository import PatientRepository
from app.modules.sessions.models import Session, SessionStatus
from app.modules.sessions.repository import SessionRepository
from app.modules.users.models import User

# ---------------------------------------------------------------------------
# HTML templates (CSS braces escaped as {{ }})
# ---------------------------------------------------------------------------

_SESSION_REPORT_TEMPLATE = """
<!DOCTYPE html>
<html lang="es">
<head>
  <meta charset="UTF-8">
  <style>
    body {{ font-family: Arial, sans-serif; font-size: 12px; color: #333; margin: 40px; }}
    h1 {{ color: #4A5568; border-bottom: 2px solid #4A5568; padding-bottom: 8px; }}
    h2 {{ color: #2D3748; margin-top: 24px; }}
    table {{ width: 100%; border-collapse: collapse; margin-top: 12px; }}
    th {{ background: #4A5568; color: white; padding: 8px; text-align: left; }}
    td {{ padding: 6px 8px; border-bottom: 1px solid #E2E8F0; }}
    tr:nth-child(even) {{ background: #F7FAFC; }}
    .footer {{ margin-top: 40px; font-size: 10px; color: #718096; text-align: center; }}
    .badge {{ display: inline-block; padding: 2px 8px; border-radius: 4px;
              background: #EBF8FF; color: #2B6CB0; font-weight: bold; }}
  </style>
</head>
<body>
  <h1>Reporte de Sesión Clínica — Serena</h1>
  <p><strong>Paciente:</strong> {patient_name} ({patient_code})</p>
  <p><strong>Terapeuta:</strong> {therapist_name}</p>
  <p><strong>Fecha de sesión:</strong> {scheduled_at}</p>
  <p><strong>Duración:</strong> {duration}</p>
  <p><strong>Estado:</strong> <span class="badge">{status}</span></p>

  <h2>Notas del Terapeuta</h2>
  <p>{notes}</p>

  <h2>Análisis Emocional — Resumen</h2>
  <table>
    <tr><th>Emoción</th><th>Promedio</th></tr>
    <tr><td>Felicidad</td><td>{avg_happiness:.1%}</td></tr>
    <tr><td>Tristeza</td><td>{avg_sadness:.1%}</td></tr>
    <tr><td>Enojo</td><td>{avg_anger:.1%}</td></tr>
    <tr><td>Miedo</td><td>{avg_fear:.1%}</td></tr>
    <tr><td>Asco</td><td>{avg_disgust:.1%}</td></tr>
    <tr><td>Sorpresa</td><td>{avg_surprise:.1%}</td></tr>
    <tr><td>Neutral</td><td>{avg_neutral:.1%}</td></tr>
  </table>
  <p><strong>Emoción dominante:</strong> {dominant_emotion} | <strong>Total snapshots:</strong> {snapshot_count}</p>

  <h2>Microexpresiones Detectadas ({micro_count})</h2>
  {micro_table}

  <div class="footer">
    Generado el {generated_at} por Serena Sistema de Análisis Biométrico.
    Documento confidencial — solo para uso clínico.
  </div>
</body>
</html>
"""

_PATIENT_REPORT_HTML = """
<!DOCTYPE html>
<html lang="es">
<head>
  <meta charset="UTF-8">
  <style>
    body {{ font-family: Arial, sans-serif; font-size: 11px; color: #1a2e35; margin: 40px; }}
    .hdr {{ background: #003441; color: white; padding: 16px 20px; margin-bottom: 20px; }}
    .hdr h1 {{ margin: 0 0 4px 0; font-size: 18px; font-weight: bold; }}
    .hdr p {{ margin: 0; font-size: 10px; }}
    .info-tbl {{ width: 100%; border-collapse: collapse; margin-bottom: 20px;
                 background: #f8fafb; border: 1px solid #e1e3e4; font-size: 11px; }}
    .info-tbl td {{ padding: 6px 12px; vertical-align: top; width: 25%; }}
    .info-tbl .lbl {{ font-size: 9px; font-weight: bold; text-transform: uppercase;
                      letter-spacing: 0.08em; color: #6b7280; display: block; margin-bottom: 1px; }}
    h2 {{ font-size: 13px; font-weight: bold; color: #003441; border-bottom: 1px solid #e1e3e4;
          padding-bottom: 5px; margin: 20px 0 10px; }}
    table.data {{ width: 100%; border-collapse: collapse; font-size: 10px; }}
    table.data th {{ background: #003441; color: white; padding: 7px 8px; text-align: left; font-size: 9px; font-weight: bold; }}
    table.data td {{ padding: 6px 8px; border-bottom: 1px solid #f0f2f3; vertical-align: top; }}
    table.data tr:nth-child(even) td {{ background: #f8fafb; }}
    .badge {{ padding: 1px 6px; font-size: 8px; font-weight: bold; text-transform: uppercase; }}
    .badge-completed {{ background: #dcfce7; color: #166534; }}
    .badge-scheduled {{ background: #dbeafe; color: #1d4ed8; }}
    .badge-cancelled {{ background: #fee2e2; color: #991b1b; }}
    .badge-active {{ background: #fef9c3; color: #854d0e; }}
    .r-pos {{ color: #166534; font-weight: bold; }}
    .r-neg {{ color: #991b1b; font-weight: bold; }}
    .notes-box {{ padding: 10px 14px; background: #f8fafb; border-left: 3px solid #003441;
                  font-size: 11px; line-height: 1.5; font-style: italic; margin-bottom: 16px; }}
    .footer {{ margin-top: 32px; padding-top: 12px; border-top: 1px solid #e1e3e4;
               font-size: 9px; color: #9ca3af; text-align: center; }}
    .num {{ text-align: right; }}
  </style>
</head>
<body>
  <div class="hdr">
    <h1>Expediente Clinico: {patient_name}</h1>
    <p>Codigo: {patient_code} | Terapeuta: {therapist_name} | Generado: {generated_at}</p>
  </div>

  <table class="info-tbl">
    <tr>
      <td><span class="lbl">Nombre Completo</span>{patient_name}</td>
      <td><span class="lbl">Codigo de Paciente</span>{patient_code}</td>
      <td><span class="lbl">Correo Electronico</span>{email}</td>
      <td><span class="lbl">Telefono</span>{phone}</td>
    </tr>
    <tr>
      <td><span class="lbl">Fecha de Nacimiento</span>{birth_date}</td>
      <td><span class="lbl">Genero</span>{gender}</td>
      <td><span class="lbl">Terapeuta Asignado</span>{therapist_name}</td>
      <td><span class="lbl">Fecha de Registro</span>{created_at}</td>
    </tr>
  </table>

  {notes_section}

  <h2>Historial de Sesiones ({session_count} sesiones)</h2>
  {sessions_table}

  {evolution_section}

  <div class="footer">
    Reporte generado el {generated_at} por SerenaMente - Sistema de Analisis Biometrico.
    Documento confidencial, solo para uso clinico interno.
  </div>
</body>
</html>
"""

_ADMIN_REPORT_HTML = """
<!DOCTYPE html>
<html lang="es">
<head>
  <meta charset="UTF-8">
  <style>
    body {{ font-family: Arial, sans-serif; font-size: 11px; color: #1a2e35; margin: 40px; }}
    .hdr {{ background: #003441; color: white; padding: 16px 20px; margin-bottom: 20px; }}
    .hdr h1 {{ margin: 0 0 4px 0; font-size: 18px; font-weight: bold; }}
    .hdr p {{ margin: 0; font-size: 10px; }}
    .stats-tbl {{ width: 100%; border-collapse: collapse; margin-bottom: 20px; }}
    .stats-tbl td {{ width: 33%; padding: 16px; background: #f8fafb; border: 1px solid #e1e3e4;
                     text-align: center; }}
    .stats-tbl .val {{ font-size: 28px; font-weight: bold; color: #003441; display: block; }}
    .stats-tbl .lbl {{ font-size: 9px; color: #6b7280; text-transform: uppercase;
                       letter-spacing: 0.08em; margin-top: 3px; display: block; }}
    h2 {{ font-size: 13px; font-weight: bold; color: #003441; border-bottom: 1px solid #e1e3e4;
          padding-bottom: 5px; margin: 20px 0 10px; }}
    table.data {{ width: 100%; border-collapse: collapse; font-size: 10px; }}
    table.data th {{ background: #003441; color: white; padding: 7px 8px; text-align: left; font-size: 9px; font-weight: bold; }}
    table.data td {{ padding: 6px 8px; border-bottom: 1px solid #f0f2f3; }}
    table.data tr:nth-child(even) td {{ background: #f8fafb; }}
    .num {{ text-align: right; }}
    .footer {{ margin-top: 32px; padding-top: 12px; border-top: 1px solid #e1e3e4;
               font-size: 9px; color: #9ca3af; text-align: center; }}
  </style>
</head>
<body>
  <div class="hdr">
    <h1>Reporte Administrativo - SerenaMente</h1>
    <p>Periodo: {date_from} al {date_to} | Generado: {generated_at}</p>
  </div>

  <table class="stats-tbl">
    <tr>
      <td><span class="val">{total_patients}</span><span class="lbl">Total Pacientes</span></td>
      <td><span class="val">{active_patients}</span><span class="lbl">Pacientes Activos</span></td>
      <td><span class="val">{new_patients}</span><span class="lbl">Nuevos en el Periodo</span></td>
    </tr>
  </table>

  <h2>Pacientes Nuevos por Mes</h2>
  {monthly_table}

  <h2>Sesiones por Terapeuta en el Periodo</h2>
  {therapist_table}

  <div class="footer">
    Reporte generado el {generated_at} por SerenaMente - Sistema de Analisis Biometrico.
    Documento confidencial, solo para uso administrativo.
  </div>
</body>
</html>
"""

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

_STATUS_LABELS: dict[str, str] = {
    "scheduled": "Programada",
    "active": "En Curso",
    "completed": "Completada",
    "cancelled": "Cancelada",
}

_STATUS_BADGE: dict[str, str] = {
    "scheduled": "badge-scheduled",
    "active": "badge-active",
    "completed": "badge-completed",
    "cancelled": "badge-cancelled",
}

_EMOTION_LABELS: list[tuple[str, str]] = [
    ("avg_happiness", "Felicidad"),
    ("avg_sadness", "Tristeza"),
    ("avg_anger", "Enojo"),
    ("avg_fear", "Miedo"),
    ("avg_disgust", "Asco"),
    ("avg_surprise", "Sorpresa"),
    ("avg_neutral", "Neutral"),
]

# ---------------------------------------------------------------------------
# Small helpers
# ---------------------------------------------------------------------------


def _status_val(session: Session) -> str:
    s = session.status
    return s.value if hasattr(s, "value") else str(s)


def _duration_str(session: Session) -> str:
    if session.started_at and session.ended_at:
        secs = int((session.ended_at - session.started_at).total_seconds())
        return f"{secs // 60}m {secs % 60}s"
    return "—"


def _format_date(d) -> str:
    if d is None:
        return "N/A"
    if hasattr(d, "strftime"):
        return d.strftime("%d/%m/%Y")
    return str(d)


def _dominant_emotion(avgs: dict) -> str:
    best_key, best_val = "avg_neutral", -1.0
    for key, _ in _EMOTION_LABELS:
        if avgs.get(key, 0.0) > best_val:
            best_val = avgs[key]
            best_key = key
    label_map = {k: lbl for k, lbl in _EMOTION_LABELS}
    return label_map.get(best_key, "Neutral")


def _ratio(avgs: dict) -> tuple[str, bool]:
    """Return (formatted_ratio, is_positive)."""
    pos = avgs.get("avg_happiness", 0) + avgs.get("avg_surprise", 0) + avgs.get("avg_neutral", 0)
    neg = avgs.get("avg_sadness", 0) + avgs.get("avg_anger", 0) + avgs.get("avg_fear", 0) + avgs.get("avg_disgust", 0)
    if neg < 0.005:
        return ("—", True) if pos < 0.005 else (">5.0", True)
    val = pos / neg
    return f"{val:.2f}", val >= 1.0


def _ratio_numeric(avgs: dict) -> float | None:
    """Return numeric ratio value, or None when no emotional expression detected."""
    pos = avgs.get("avg_happiness", 0) + avgs.get("avg_surprise", 0) + avgs.get("avg_neutral", 0)
    neg = avgs.get("avg_sadness", 0) + avgs.get("avg_anger", 0) + avgs.get("avg_fear", 0) + avgs.get("avg_disgust", 0)
    if neg < 0.005:
        return None if pos < 0.005 else 5.0
    return round(pos / neg, 4)


def _build_ratio_svg(evolution: list[tuple]) -> str:
    """Generate an inline SVG line chart of the ratio evolution for embedding in PDF HTML."""
    if not evolution:
        return ""

    W, H = 520, 180
    PL, PR, PT, PB = 44, 32, 24, 42
    plot_w = W - PL - PR
    plot_h = H - PT - PB
    bottom = PT + plot_h

    # Compute per-session values
    data: list[tuple[float, str]] = []
    for session, avgs in evolution:
        r = _ratio_numeric(avgs)
        d = session.scheduled_at
        date_lbl = f"{d.day:02d}/{d.month:02d}" if d else ""
        data.append((r if r is not None else 0.0, date_lbl))

    ratio_vals = [r for r, _ in data]
    max_r = max(ratio_vals + [2.0])
    max_y = round(((max_r + 0.5) * 2) // 1 / 2 + 0.5, 1)

    step = 0.5 if max_y <= 3.5 else (1.0 if max_y <= 7.0 else 2.0)

    # Y-axis labels
    y_labels: list[tuple[float, float]] = []
    v = 0.0
    while v <= max_y + 0.01:
        y = round(PT + plot_h * (1.0 - v / max_y), 1)
        y_labels.append((round(v, 1), y))
        v = round(v + step, 2)

    balance_y = round(PT + plot_h * (1.0 - 1.0 / max_y), 1)

    n = len(data)
    pts = []
    for i, (r, date_lbl) in enumerate(data):
        x = round(PL + (i / (n - 1) * plot_w), 1) if n > 1 else round(PL + plot_w / 2, 1)
        y = round(PT + plot_h * (1.0 - min(r, max_y) / max_y), 1)
        pts.append({"x": x, "y": y, "r": r, "pos": r >= 1.0,
                    "lbl": f"{r:.2f}" if r > 0 else "—",
                    "date": date_lbl, "idx": i + 1})

    parts: list[str] = []

    # Grid lines + Y labels
    for v, y in y_labels:
        parts.append(f'<line x1="{PL}" x2="{W - PR}" y1="{y}" y2="{y}" stroke="#f0f0f0" stroke-width="1"/>')
        parts.append(f'<text x="{PL - 3}" y="{y + 3}" text-anchor="end" font-size="8" fill="#aaa">{v:.1f}</text>')

    # Balance line at ratio = 1.0
    parts.append(
        f'<line x1="{PL}" x2="{W - PR}" y1="{balance_y}" y2="{balance_y}" '
        f'stroke="#d97706" stroke-width="1.5" stroke-dasharray="4,2" opacity="0.9"/>'
    )
    parts.append(
        f'<text x="{W - PR + 2}" y="{balance_y + 3}" font-size="7" fill="#d97706" font-weight="bold">1.0</text>'
    )

    # Area fill + line (only meaningful with 2+ points)
    if n > 1:
        area_d = (f"M{pts[0]['x']},{bottom} " +
                  " ".join(f"L{p['x']},{p['y']}" for p in pts) +
                  f" L{pts[-1]['x']},{bottom} Z")
        parts.append(f'<path d="{area_d}" fill="#003441" fill-opacity="0.08"/>')
        line_d = " ".join(f"{'M' if i == 0 else 'L'}{p['x']},{p['y']}" for i, p in enumerate(pts))
        parts.append(
            f'<path d="{line_d}" fill="none" stroke="#003441" stroke-width="2" '
            f'stroke-linejoin="round" stroke-linecap="round"/>'
        )

    # Dots, value labels, X labels
    for p in pts:
        dot_color = "#22c55e" if p["pos"] else "#ef4444"
        lbl_color = "#166534" if p["pos"] else "#991b1b"
        lbl_y = max(p["y"] - 8, PT - 2)
        parts.append(
            f'<text x="{p["x"]}" y="{lbl_y}" text-anchor="middle" '
            f'font-size="8" font-weight="bold" fill="{lbl_color}">{p["lbl"]}</text>'
        )
        parts.append(
            f'<circle cx="{p["x"]}" cy="{p["y"]}" r="4" '
            f'fill="{dot_color}" stroke="white" stroke-width="2"/>'
        )
        parts.append(
            f'<text x="{p["x"]}" y="{bottom + 12}" text-anchor="middle" '
            f'font-size="8" fill="#333" font-weight="600">S{p["idx"]}</text>'
        )
        parts.append(
            f'<text x="{p["x"]}" y="{bottom + 22}" text-anchor="middle" '
            f'font-size="7" fill="#aaa">{p["date"]}</text>'
        )

    # Axes
    parts.append(f'<line x1="{PL}" x2="{PL}" y1="{PT}" y2="{bottom}" stroke="#ddd" stroke-width="1"/>')
    parts.append(f'<line x1="{PL}" x2="{W - PR}" y1="{bottom}" y2="{bottom}" stroke="#ddd" stroke-width="1"/>')

    inner = "\n  ".join(parts)
    return (
        f'<svg viewBox="0 0 {W} {H}" xmlns="http://www.w3.org/2000/svg" '
        f'style="width:100%;height:auto;display:block;margin-bottom:12px">\n  {inner}\n</svg>'
    )


def _xl_header(ws, row: int, headers: list[str], fill_color: str = "003441") -> None:
    from openpyxl.styles import Font, PatternFill, Alignment

    fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    font = Font(bold=True, color="FFFFFF", size=10)
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col, text in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=text)
        cell.font = font
        cell.fill = fill
        cell.alignment = align


def _xl_auto_width(ws) -> None:
    from openpyxl.utils import get_column_letter

    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 3, 55)


# ---------------------------------------------------------------------------
# Service
# ---------------------------------------------------------------------------


class ReportService:
    """Business logic for generating clinical PDF and Excel reports."""

    def __init__(self, db: AsyncSession) -> None:
        self._db = db
        self._session_repo = SessionRepository(db)
        self._patient_repo = PatientRepository(db)
        self._snapshot_repo = EmotionalSnapshotRepository(db)
        self._micro_repo = MicroexpressionRepository(db)

    # ------------------------------------------------------------------
    # Existing: session PDF
    # ------------------------------------------------------------------

    async def generate_session_pdf(self, session_id: uuid.UUID) -> bytes:
        from weasyprint import HTML

        session = await self._session_repo.get_by_id(session_id)
        if not session:
            raise NotFoundError("Session")

        averages = await self._snapshot_repo.get_session_averages(session_id)
        microexpressions = await self._micro_repo.list_by_session(session_id)

        duration = "N/A"
        if session.started_at and session.ended_at:
            secs = int((session.ended_at - session.started_at).total_seconds())
            duration = f"{secs // 60}m {secs % 60}s"

        micro_rows = "".join(
            f"<tr><td>{m.timestamp_offset:.1f}s</td>"
            f"<td>{m.emotion_detected}</td>"
            f"<td>{m.intensity:.2f}</td>"
            f"<td>{m.duration_ms}ms</td></tr>"
            for m in microexpressions
        )
        micro_table = (
            f"<table><tr><th>Tiempo</th><th>Emoción</th><th>Intensidad</th><th>Duración</th></tr>"
            f"{micro_rows}</table>"
            if microexpressions
            else "<p>No se detectaron microexpresiones en esta sesión.</p>"
        )

        html_content = _SESSION_REPORT_TEMPLATE.format(
            patient_name=session.patient.full_name,
            patient_code=session.patient.code,
            therapist_name=session.therapist.full_name,
            scheduled_at=session.scheduled_at.strftime("%d/%m/%Y %H:%M"),
            duration=duration,
            status=session.status.value.upper(),
            notes=session.notes or "Sin notas registradas.",
            avg_happiness=averages["avg_happiness"],
            avg_sadness=averages["avg_sadness"],
            avg_anger=averages["avg_anger"],
            avg_fear=averages["avg_fear"],
            avg_disgust=averages["avg_disgust"],
            avg_surprise=averages["avg_surprise"],
            avg_neutral=averages["avg_neutral"],
            dominant_emotion=averages["dominant_overall"].upper(),
            snapshot_count=averages["snapshot_count"],
            micro_count=len(microexpressions),
            micro_table=micro_table,
            generated_at=datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
        )

        logger.info("Generating PDF report | session={}", session_id)
        return HTML(string=html_content, base_url=None).write_pdf(presentational_hints=True)

    # ------------------------------------------------------------------
    # Patient PDF
    # ------------------------------------------------------------------

    async def generate_patient_pdf(self, patient_id: uuid.UUID) -> bytes:
        from weasyprint import HTML

        patient, sessions, therapist_name, evolution = await self._gather_patient_data(patient_id)

        notes_section = (
            f'<h2>Notas Clínicas</h2><div class="notes-box">{patient.medical_notes}</div>'
            if patient.medical_notes
            else ""
        )

        # Sessions table
        if sessions:
            avgs_map = {s.id: e for s, e in evolution}
            header = "<tr><th>Fecha y Hora</th><th>Estado</th><th>Duración</th><th>Notas</th><th>Emoción Dominante</th></tr>"
            rows = ""
            for s in sessions:
                sv = _status_val(s)
                avgs = avgs_map.get(s.id)
                dominant = _dominant_emotion(avgs) if avgs and avgs.get("snapshot_count", 0) > 0 else "—"
                note = (s.notes or "")[:120] + ("…" if len(s.notes or "") > 120 else "")
                date_str = s.scheduled_at.strftime("%d/%m/%Y %H:%M") if s.scheduled_at else "—"
                badge_cls = _STATUS_BADGE.get(sv, "badge-scheduled")
                status_lbl = _STATUS_LABELS.get(sv, sv)
                rows += (
                    f"<tr><td>{date_str}</td>"
                    f"<td><span class='badge {badge_cls}'>{status_lbl}</span></td>"
                    f"<td>{_duration_str(s)}</td>"
                    f"<td>{note or '—'}</td>"
                    f"<td>{dominant}</td></tr>"
                )
            sessions_table = f'<table class="data">{header}{rows}</table>'
        else:
            sessions_table = "<p>No hay sesiones registradas para este paciente.</p>"

        # Evolution table
        evolution_section = ""
        if evolution:
            evo_header = (
                "<tr><th>Sesión</th><th>Fecha</th><th>Ratio P/N</th>"
                "<th>Alegría</th><th>Tristeza</th><th>Enojo</th>"
                "<th>Miedo</th><th>Asco</th><th>Sorpresa</th><th>Neutral</th></tr>"
            )
            evo_rows = ""
            for idx, (s, avgs) in enumerate(evolution, start=1):
                ratio_str, is_pos = _ratio(avgs)
                cls = "r-pos" if is_pos else "r-neg"
                date_str = s.scheduled_at.strftime("%d/%m/%Y") if s.scheduled_at else "—"
                evo_rows += (
                    f"<tr><td>S{idx}</td><td>{date_str}</td>"
                    f"<td class='{cls}'>{ratio_str}</td>"
                    f"<td>{avgs['avg_happiness']:.1%}</td>"
                    f"<td>{avgs['avg_sadness']:.1%}</td>"
                    f"<td>{avgs['avg_anger']:.1%}</td>"
                    f"<td>{avgs['avg_fear']:.1%}</td>"
                    f"<td>{avgs['avg_disgust']:.1%}</td>"
                    f"<td>{avgs['avg_surprise']:.1%}</td>"
                    f"<td>{avgs['avg_neutral']:.1%}</td></tr>"
                )
            evolution_section = (
                f"<h2>Evolucion del Ratio de Activacion Emocional</h2>"
                + _build_ratio_svg(evolution)
                + f'<table class="data">{evo_header}{evo_rows}</table>'
            )

        html = _PATIENT_REPORT_HTML.format(
            patient_name=patient.full_name,
            patient_code=patient.code,
            therapist_name=therapist_name,
            email=patient.email or "No registrado",
            phone=patient.phone or "No registrado",
            birth_date=_format_date(patient.birth_date),
            gender=(patient.gender or "N/A").capitalize(),
            created_at=_format_date(patient.created_at),
            session_count=len(sessions),
            notes_section=notes_section,
            sessions_table=sessions_table,
            evolution_section=evolution_section,
            generated_at=datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
        )

        logger.info("Generating patient PDF | patient={}", patient_id)
        return HTML(string=html, base_url=None).write_pdf(presentational_hints=True)

    # ------------------------------------------------------------------
    # Patient Excel
    # ------------------------------------------------------------------

    async def generate_patient_excel(self, patient_id: uuid.UUID) -> bytes:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill

        patient, sessions, therapist_name, evolution = await self._gather_patient_data(patient_id)

        wb = Workbook()

        # Sheet 1 — Patient info
        ws1 = wb.active
        ws1.title = "Paciente"
        info_rows = [
            ("Nombre Completo", patient.full_name),
            ("Código", patient.code),
            ("Terapeuta Asignado", therapist_name),
            ("Correo Electrónico", patient.email or "No registrado"),
            ("Teléfono", patient.phone or "No registrado"),
            ("Fecha de Nacimiento", _format_date(patient.birth_date)),
            ("Género", (patient.gender or "N/A").capitalize()),
            ("Fecha de Registro", _format_date(patient.created_at)),
            ("Notas Clínicas", patient.medical_notes or "Sin notas"),
        ]
        label_font = Font(bold=True, color="003441", size=10)
        for r, (label, value) in enumerate(info_rows, start=1):
            ws1.cell(row=r, column=1, value=label).font = label_font
            ws1.cell(row=r, column=2, value=value)
        ws1.column_dimensions["A"].width = 25
        ws1.column_dimensions["B"].width = 50

        # Sheet 2 — Sessions
        ws2 = wb.create_sheet("Sesiones")
        headers2 = ["Fecha y Hora", "Estado", "Duración", "Emoción Dominante",
                    "Alegría", "Tristeza", "Enojo", "Miedo", "Asco", "Sorpresa", "Neutral", "Notas"]
        _xl_header(ws2, 1, headers2)
        avgs_map = {s.id: e for s, e in evolution}
        for r, s in enumerate(sessions, start=2):
            avgs = avgs_map.get(s.id)
            has_bio = avgs and avgs.get("snapshot_count", 0) > 0
            sv = _status_val(s)
            ws2.cell(row=r, column=1, value=s.scheduled_at.strftime("%d/%m/%Y %H:%M") if s.scheduled_at else "—")
            ws2.cell(row=r, column=2, value=_STATUS_LABELS.get(sv, sv))
            ws2.cell(row=r, column=3, value=_duration_str(s))
            ws2.cell(row=r, column=4, value=_dominant_emotion(avgs) if has_bio else "—")
            for ci, (key, _) in enumerate(_EMOTION_LABELS, start=5):
                ws2.cell(row=r, column=ci, value=round(avgs.get(key, 0), 4) if has_bio else None)
            ws2.cell(row=r, column=12, value=s.notes or "")
        _xl_auto_width(ws2)

        # Sheet 3 — Evolution
        ws3 = wb.create_sheet("Evolución Emocional")
        headers3 = ["Sesión", "Fecha", "Ratio P/N",
                    "Alegría", "Tristeza", "Enojo", "Miedo", "Asco", "Sorpresa", "Neutral"]
        _xl_header(ws3, 1, headers3)
        for idx, (s, avgs) in enumerate(evolution, start=1):
            ws3.cell(row=idx + 1, column=1, value=f"S{idx}")
            ws3.cell(row=idx + 1, column=2, value=s.scheduled_at.strftime("%d/%m/%Y") if s.scheduled_at else "—")
            ws3.cell(row=idx + 1, column=3, value=_ratio_numeric(avgs))  # numeric for chart
            for ci, (key, _) in enumerate(_EMOTION_LABELS, start=4):
                ws3.cell(row=idx + 1, column=ci, value=round(avgs.get(key, 0), 4))
        _xl_auto_width(ws3)

        # Line chart of ratio evolution
        if len(evolution) >= 2:
            from openpyxl.chart import LineChart, Reference

            chart = LineChart()
            chart.title = "Ratio de Activacion Emocional"
            chart.style = 10
            chart.y_axis.title = "Ratio P/N"
            chart.x_axis.title = "Sesion"
            chart.y_axis.numFmt = "0.00"
            chart.width = 22
            chart.height = 13

            data_ref = Reference(ws3, min_col=3, min_row=1, max_row=len(evolution) + 1)
            chart.add_data(data_ref, titles_from_data=True)
            cats = Reference(ws3, min_col=1, min_row=2, max_row=len(evolution) + 1)
            chart.set_categories(cats)

            # Style: teal line with markers
            series = chart.series[0]
            series.graphicalProperties.line.solidFill = "003441"
            series.graphicalProperties.line.width = 25000  # 2.5pt in EMU
            series.marker.symbol = "circle"
            series.marker.size = 6
            series.marker.graphicalProperties.fgColor = "003441"
            series.marker.graphicalProperties.line.solidFill = "003441"

            ws3.add_chart(chart, f"A{len(evolution) + 4}")

        buf = BytesIO()
        wb.save(buf)
        logger.info("Generating patient Excel | patient={}", patient_id)
        return buf.getvalue()

    # ------------------------------------------------------------------
    # Admin PDF
    # ------------------------------------------------------------------

    async def generate_admin_pdf(self, date_from: date, date_to: date) -> bytes:
        from weasyprint import HTML

        totals, monthly_rows, therapist_rows = await self._gather_admin_data(date_from, date_to)

        if monthly_rows:
            m_header = "<tr><th>Mes</th><th class='num'>Nuevos Pacientes</th></tr>"
            m_rows = "".join(
                f"<tr><td>{row[0]}</td><td class='num'>{row[1]}</td></tr>"
                for row in monthly_rows
            )
            monthly_table = f'<table class="data">{m_header}{m_rows}</table>'
        else:
            monthly_table = "<p>No se registraron pacientes nuevos en el período indicado.</p>"

        if therapist_rows:
            t_header = (
                "<tr><th>Terapeuta</th><th class='num'>Total</th>"
                "<th class='num'>Completadas</th><th class='num'>Programadas</th>"
                "<th class='num'>Canceladas</th></tr>"
            )
            t_rows = "".join(
                f"<tr><td>{r[0]}</td><td class='num'>{r[1]}</td>"
                f"<td class='num'>{r[2]}</td><td class='num'>{r[3]}</td>"
                f"<td class='num'>{r[4]}</td></tr>"
                for r in therapist_rows
            )
            therapist_table = f'<table class="data">{t_header}{t_rows}</table>'
        else:
            therapist_table = "<p>No hay sesiones registradas en el período indicado.</p>"

        html = _ADMIN_REPORT_HTML.format(
            date_from=date_from.strftime("%d/%m/%Y"),
            date_to=date_to.strftime("%d/%m/%Y"),
            total_patients=totals["total"],
            active_patients=totals["active"],
            new_patients=totals["new"],
            monthly_table=monthly_table,
            therapist_table=therapist_table,
            generated_at=datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
        )

        logger.info("Generating admin PDF | {} to {}", date_from, date_to)
        return HTML(string=html, base_url=None).write_pdf(presentational_hints=True)

    # ------------------------------------------------------------------
    # Admin Excel
    # ------------------------------------------------------------------

    async def generate_admin_excel(self, date_from: date, date_to: date) -> bytes:
        from openpyxl import Workbook
        from openpyxl.styles import Font

        totals, monthly_rows, therapist_rows = await self._gather_admin_data(date_from, date_to)

        wb = Workbook()

        # Sheet 1 — Summary
        ws1 = wb.active
        ws1.title = "Resumen"
        summary_data = [
            ("Período", f"{date_from.strftime('%d/%m/%Y')} — {date_to.strftime('%d/%m/%Y')}"),
            ("Total de Pacientes Registrados", totals["total"]),
            ("Pacientes Activos", totals["active"]),
            ("Pacientes Nuevos en el Período", totals["new"]),
            ("Generado el", datetime.now().strftime("%d/%m/%Y %H:%M:%S")),
        ]
        label_font = Font(bold=True, color="003441", size=10)
        for r, (label, value) in enumerate(summary_data, start=1):
            ws1.cell(row=r, column=1, value=label).font = label_font
            ws1.cell(row=r, column=2, value=value)
        ws1.column_dimensions["A"].width = 40
        ws1.column_dimensions["B"].width = 35

        # Sheet 2 — Monthly new patients
        ws2 = wb.create_sheet("Nuevos por Mes")
        _xl_header(ws2, 1, ["Mes", "Pacientes Nuevos"])
        for r, (month_str, count) in enumerate(monthly_rows, start=2):
            ws2.cell(row=r, column=1, value=month_str)
            ws2.cell(row=r, column=2, value=count)
        _xl_auto_width(ws2)

        # Sheet 3 — Sessions by therapist
        ws3 = wb.create_sheet("Por Terapeuta")
        _xl_header(ws3, 1, ["Terapeuta", "Total", "Completadas", "Programadas", "Canceladas"])
        for r, (name, total, completed, scheduled, cancelled) in enumerate(therapist_rows, start=2):
            ws3.cell(row=r, column=1, value=name)
            ws3.cell(row=r, column=2, value=total)
            ws3.cell(row=r, column=3, value=completed)
            ws3.cell(row=r, column=4, value=scheduled)
            ws3.cell(row=r, column=5, value=cancelled)
        _xl_auto_width(ws3)

        buf = BytesIO()
        wb.save(buf)
        logger.info("Generating admin Excel | {} to {}", date_from, date_to)
        return buf.getvalue()

    # ------------------------------------------------------------------
    # Private data-gathering helpers
    # ------------------------------------------------------------------

    async def _gather_patient_data(
        self, patient_id: uuid.UUID
    ) -> tuple[Patient, list[Session], str, list[tuple[Session, dict]]]:
        """Fetch patient, sessions, therapist name, and evolution data."""
        patient = await self._patient_repo.get_by_id(patient_id)
        if not patient:
            raise NotFoundError("Patient")

        # Resolve therapist name without relying on lazy-loaded relationship
        therapist_name = "No asignado"
        if patient.therapist_id:
            stmt = select(User).where(User.id == patient.therapist_id)
            result = await self._db.execute(stmt)
            user = result.scalar_one_or_none()
            if user:
                therapist_name = user.full_name

        # Fetch all sessions for this patient, oldest first
        raw_sessions, _ = await self._session_repo.list_paginated(
            page=1, page_size=500, patient_id=patient_id
        )
        sessions: list[Session] = list(reversed(raw_sessions))

        # Gather biometric averages for completed sessions that have data
        evolution: list[tuple[Session, dict]] = []
        for s in sessions:
            sv = _status_val(s)
            if sv == "completed":
                avgs = await self._snapshot_repo.get_session_averages(s.id)
                if avgs.get("snapshot_count", 0) > 0:
                    evolution.append((s, avgs))

        return patient, sessions, therapist_name, evolution

    async def _gather_admin_data(
        self, date_from: date, date_to: date
    ) -> tuple[dict, list[tuple], list[tuple]]:
        """Run the three admin aggregation queries."""

        # --- Patient totals ---
        total = (await self._db.execute(
            select(func.count(Patient.id)).where(Patient.deleted_at.is_(None))
        )).scalar() or 0

        active = (await self._db.execute(
            select(func.count(Patient.id)).where(
                and_(Patient.deleted_at.is_(None), Patient.is_active.is_(True))
            )
        )).scalar() or 0

        new_in_period = (await self._db.execute(
            select(func.count(Patient.id)).where(
                and_(
                    Patient.deleted_at.is_(None),
                    func.date(Patient.created_at) >= date_from,
                    func.date(Patient.created_at) <= date_to,
                )
            )
        )).scalar() or 0

        # --- Monthly breakdown ---
        month_col = func.date_trunc("month", Patient.created_at).label("month")
        monthly_result = (await self._db.execute(
            select(month_col, func.count(Patient.id).label("count"))
            .where(
                and_(
                    Patient.deleted_at.is_(None),
                    func.date(Patient.created_at) >= date_from,
                    func.date(Patient.created_at) <= date_to,
                )
            )
            .group_by(month_col)
            .order_by(month_col)
        )).all()

        # Format month labels
        month_map = {1: "Ene", 2: "Feb", 3: "Mar", 4: "Abr", 5: "May", 6: "Jun",
                     7: "Jul", 8: "Ago", 9: "Sep", 10: "Oct", 11: "Nov", 12: "Dic"}
        monthly_rows = [
            (
                f"{month_map.get(row.month.month, '?')} {row.month.year}" if row.month else "—",
                row.count,
            )
            for row in monthly_result
        ]

        # --- Sessions by therapist ---
        therapist_result = (await self._db.execute(
            select(
                User.full_name.label("therapist"),
                func.count(Session.id).label("total"),
                func.count(Session.id).filter(Session.status == SessionStatus.COMPLETED).label("completed"),
                func.count(Session.id).filter(Session.status == SessionStatus.SCHEDULED).label("scheduled"),
                func.count(Session.id).filter(Session.status == SessionStatus.CANCELLED).label("cancelled"),
            )
            .select_from(Session)
            .join(User, User.id == Session.therapist_id)
            .where(
                and_(
                    User.deleted_at.is_(None),
                    func.date(Session.scheduled_at) >= date_from,
                    func.date(Session.scheduled_at) <= date_to,
                )
            )
            .group_by(User.id, User.full_name)
            .order_by(User.full_name)
        )).all()

        therapist_rows = [
            (r.therapist, r.total, r.completed, r.scheduled, r.cancelled)
            for r in therapist_result
        ]

        return (
            {"total": total, "active": active, "new": new_in_period},
            monthly_rows,
            therapist_rows,
        )
